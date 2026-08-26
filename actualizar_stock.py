# -*- coding: utf-8 -*-
"""
Actualiza el stock ("Stock AP") del listado maestro de obsolescencia a partir
del archivo diario de stock SAP ("Libre utilizacion") que se deja en la
carpeta "Actualizacion de Stock/" con el nombre Stock_DD_MM.xlsx.

Reglas de negocio (confirmadas con el usuario el 2026-08-06):
- Solo se considera el almacen '1100' de cada tienda; otros codigos de
  almacen (ej. 1600) se ignoran por completo.
- Si una tienda entera no aparece en el archivo del dia (ningun almacen),
  su stock NO se toca -- se asume un problema de exportacion, no falta de
  stock real -- y se reporta como "tienda ausente".
- Los productos que aparecen en el archivo del dia pero no existen en el
  listado maestro se ignoran (no se agregan: falta precio, marca, categoria
  y aplicacion vehicular) y se reportan aparte para revision manual --
  EXCEPTO los materiales de MATERIALES_NUEVOS (ver materiales_nuevos.py):
  para esos ya se conoce todo el dato fijo (llegaron por CD), asi que se
  agrega una fila nueva por tienda con esos mismos datos, la Zona real de
  esa tienda (tomada de cualquier otra fila existente de esa tienda) y el
  stock del dia. Confirmado con el usuario el 2026-08-26.
- Los productos que quedan en stock 0 NO se eliminan del listado: el portal
  los muestra marcados como "Agotado" (ver plantilla_portal.html).
- El stock NUNCA sube ni se "recupera" solo (confirmado con el usuario el
  2026-08-17, tras un problema causado por productos que volvieron a
  aparecer con stock): son productos en liquidacion, asi que solo se
  aplican bajas (incluido llegar a 0 = Agotado). Si el archivo diario trae
  mas stock que el maestro para una clave, se ignora esa fila y se reporta
  aparte en vez de aplicarse.

El cruce Centro/Almacen -> Zona/Tienda se hace por nombre: la columna
"Nombre 1" del archivo diario trae exactamente la misma etiqueta de tienda
que usa el listado maestro (ej. "AP0001-La Florida"), excepto el Centro de
Distribucion (Centro '0714'), cuyo "Nombre 1" es un nombre de operador
("Ega-Kat") y no una tienda -- para ese caso se usa el codigo de tienda fijo
"0714", que es como aparece en el maestro bajo la zona "CD".

Ejecutar desde esta carpeta:
    python actualizar_stock.py            # usa el archivo de hoy
    python actualizar_stock.py 06 08      # fuerza dia/mes (para pruebas)
"""
from __future__ import annotations

import shutil
import sys
import unicodedata
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path

import openpyxl

from materiales_nuevos import MATERIALES_NUEVOS

RAIZ_PROYECTO = Path(__file__).resolve().parent.parent
CARPETA_STOCK_DIARIO = RAIZ_PROYECTO / "Actualizacion de Stock"
MASTER = RAIZ_PROYECTO / "dashboard-obsolescencia" / "data" / "Listado Obsolecencia.xlsx"
CARPETA_RESPALDOS = CARPETA_STOCK_DIARIO / "respaldos"

CENTRO_CD = "0714"
TIENDA_CD = "0714"
ALMACEN_VALIDO = "1100"

MAX_EJEMPLOS_REPORTE = 20


def _sin_tildes(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    return "".join(c for c in s if not unicodedata.combining(c))


def _material_a_texto(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and valor == int(valor):
        return str(int(valor))
    if isinstance(valor, int):
        return str(valor)
    return str(valor).strip()


def _tienda_key(centro, nombre1) -> str:
    centro = str(centro or "").strip()
    if centro == CENTRO_CD:
        return TIENDA_CD
    return str(nombre1 or "").strip()


def localizar_archivo_del_dia(fecha: date) -> Path:
    nombre = f"Stock_{fecha:%d}_{fecha:%m}"
    candidatos = [
        p for p in CARPETA_STOCK_DIARIO.glob(f"{nombre}.*")
        if not p.name.startswith("~$")
    ]
    if not candidatos:
        raise FileNotFoundError(
            f"No encontré el archivo de stock de hoy ({nombre}.xlsx) en "
            f"«{CARPETA_STOCK_DIARIO}»."
        )
    return candidatos[0]


@dataclass
class ResultadoActualizacion:
    archivo_usado: Path
    tiendas_cubiertas: int = 0
    tiendas_ausentes: list = field(default_factory=list)
    productos_actualizados: int = 0
    productos_bajaron: list = field(default_factory=list)
    productos_agotados_nuevos: list = field(default_factory=list)
    productos_subida_ignorada: list = field(default_factory=list)
    productos_nuevos_ignorados: list = field(default_factory=list)
    filas_agregadas: list = field(default_factory=list)
    respaldo: Path | None = None


def _col(indice: dict, *nombres: str) -> int:
    for h, i in indice.items():
        for n in nombres:
            if _sin_tildes(h).lower().strip() == n:
                return i
    raise KeyError(f"No encontré ninguna columna de {nombres} en el archivo diario")


def leer_stock_diario(ruta: Path):
    """Devuelve (stock_por_clave, tiendas_presentes, texto_por_clave)."""
    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    filas = ws.iter_rows(values_only=True)
    header = next(filas)
    indice = {str(h).strip(): i for i, h in enumerate(header) if h}

    i_material = _col(indice, "material")
    i_libre = _col(indice, "libre utilizacion")
    i_centro = _col(indice, "centro")
    i_almacen = _col(indice, "almacen")
    i_nombre1 = _col(indice, "nombre 1")
    i_texto = _col(indice, "texto breve de material")

    stock: dict[tuple[str, str], float] = {}
    tiendas_presentes: set[str] = set()
    textos: dict[tuple[str, str], str] = {}

    for r in filas:
        centro, nombre1 = r[i_centro], r[i_nombre1]
        if not str(centro or "").strip() and not str(nombre1 or "").strip():
            continue
        tienda = _tienda_key(centro, nombre1)
        if not tienda:
            continue
        tiendas_presentes.add(tienda)

        almacen = str(r[i_almacen] or "").strip()
        if almacen != ALMACEN_VALIDO:
            continue
        material = _material_a_texto(r[i_material])
        if not material:
            continue
        cantidad = r[i_libre]
        cantidad = float(cantidad) if isinstance(cantidad, (int, float)) else 0
        clave = (tienda, material)
        stock[clave] = stock.get(clave, 0) + cantidad
        textos[clave] = r[i_texto]

    return stock, tiendas_presentes, textos


def actualizar(fecha: date | None = None, guardar: bool = True) -> ResultadoActualizacion:
    fecha = fecha or date.today()
    archivo_dia = localizar_archivo_del_dia(fecha)
    stock_diario, tiendas_presentes, textos_diario = leer_stock_diario(archivo_dia)

    wb = openpyxl.load_workbook(MASTER)
    ws = wb["Base"]
    filas = list(ws.iter_rows(min_row=1))
    header = [c.value for c in filas[0]]
    idx_tienda = header.index("Tienda")
    idx_zona = header.index("Zona")
    idx_material = header.index("Material")
    idx_stock = header.index("Stock AP")

    resultado = ResultadoActualizacion(archivo_usado=archivo_dia)
    resultado.tiendas_cubiertas = len(tiendas_presentes)

    tiendas_en_maestro: set[str] = set()
    vistos_hoy: set[tuple[str, str]] = set()

    # Para poder agregar filas nuevas de MATERIALES_NUEVOS en otras tiendas:
    # una plantilla de valores fijos por material (precio, marca, subcategoria,
    # aplicacion vehicular...) y la Zona real de cada tienda -- ambas se toman
    # de filas YA existentes en el maestro, nunca inventadas.
    plantilla_por_material: dict[str, list] = {}
    zona_por_tienda: dict[str, str] = {}

    for fila in filas[1:]:
        tienda = str(fila[idx_tienda].value or "").strip()
        tiendas_en_maestro.add(tienda)
        if tienda and tienda not in zona_por_tienda:
            zona_por_tienda[tienda] = fila[idx_zona].value
        material = _material_a_texto(fila[idx_material].value)
        if material in MATERIALES_NUEVOS and (
            material not in plantilla_por_material or tienda == TIENDA_CD
        ):
            plantilla_por_material[material] = [c.value for c in fila]
        if tienda not in tiendas_presentes:
            continue  # tienda ausente hoy: no se toca su stock
        clave = (tienda, material)
        vistos_hoy.add(clave)

        # SAP a veces trae "Libre utilizacion" negativa (ajustes de inventario).
        # Para el portal eso equivale a "sin stock": nunca se escribe un
        # numero negativo (se descartaria del listado en vez de marcarse
        # como Agotado).
        nuevo = max(0, stock_diario.get(clave, 0))
        anterior = fila[idx_stock].value or 0
        if nuevo == anterior:
            continue
        if nuevo > anterior:
            # Regla de negocio: el stock nunca sube ni se recupera solo.
            resultado.productos_subida_ignorada.append((tienda, material, anterior, nuevo))
            continue

        fila[idx_stock].value = nuevo
        resultado.productos_actualizados += 1
        registro = (tienda, material, anterior, nuevo)
        if nuevo == 0:
            resultado.productos_agotados_nuevos.append(registro)
        else:
            resultado.productos_bajaron.append(registro)

    resultado.tiendas_ausentes = sorted(t for t in tiendas_en_maestro if t not in tiendas_presentes)

    for clave, cantidad in stock_diario.items():
        if clave in vistos_hoy:
            continue
        tienda, material = clave
        texto = textos_diario.get(clave, "")

        # Disponibilidad real por tienda de un material ya conocido (llego
        # por CD, ver MATERIALES_NUEVOS): se agrega una fila nueva copiando
        # los datos fijos del material y la Zona real de esa tienda, en vez
        # de descartarlo. Nunca se inventa Zona: si la tienda no aparece en
        # ninguna otra fila del maestro, se reporta igual que un producto
        # nuevo (no hay de donde tomar su zona).
        plantilla = plantilla_por_material.get(material)
        zona = zona_por_tienda.get(tienda)
        cantidad_valida = max(0, cantidad)
        if plantilla is not None and zona and cantidad_valida > 0:
            nueva_fila = list(plantilla)
            nueva_fila[idx_tienda] = tienda
            nueva_fila[idx_zona] = zona
            nueva_fila[idx_material] = plantilla[idx_material]
            nueva_fila[idx_stock] = cantidad_valida
            ws.append(nueva_fila)
            resultado.filas_agregadas.append((tienda, material, texto, cantidad_valida, zona))
            continue

        resultado.productos_nuevos_ignorados.append((tienda, material, texto, cantidad))

    if guardar:
        CARPETA_RESPALDOS.mkdir(exist_ok=True)
        respaldo = CARPETA_RESPALDOS / f"Listado Obsolecencia (antes de {archivo_dia.stem}).xlsx"
        shutil.copy2(MASTER, respaldo)
        resultado.respaldo = respaldo
        wb.save(MASTER)

    return resultado


def imprimir_reporte(r: ResultadoActualizacion) -> None:
    print(f"Archivo usado: {r.archivo_usado.name}")
    if r.respaldo:
        print(f"Respaldo del listado anterior: {r.respaldo.name}")
    print(f"Tiendas cubiertas hoy: {r.tiendas_cubiertas}")
    print(f"Tiendas ausentes hoy (NO tocadas): {len(r.tiendas_ausentes)}")
    for t in r.tiendas_ausentes:
        print(f"    - {t}")
    print(f"Productos actualizados: {r.productos_actualizados}")
    print(f"  Bajaron de stock: {len(r.productos_bajaron)}")
    print(f"  Pasaron a Agotado (de >0 a 0): {len(r.productos_agotados_nuevos)}")
    for t, m, a, n in r.productos_agotados_nuevos[:MAX_EJEMPLOS_REPORTE]:
        print(f"    - {m} en {t}: {a} -> {n}")
    print(f"Subidas de stock ignoradas (regla: el stock no se recupera solo): {len(r.productos_subida_ignorada)}")
    for t, m, a, n in r.productos_subida_ignorada[:MAX_EJEMPLOS_REPORTE]:
        print(f"    - {m} en {t}: {a} -> {n} (se mantiene en {a})")
    print(f"Filas nuevas agregadas (disponibilidad por tienda de productos nuevos): {len(r.filas_agregadas)}")
    for t, m, texto, cant, zona in r.filas_agregadas[:MAX_EJEMPLOS_REPORTE]:
        print(f"    - {m} ({texto}) en {t} [{zona}]: stock {cant}")
    print(f"Productos del archivo de hoy ignorados (no estaban en el listado): {len(r.productos_nuevos_ignorados)}")
    for t, m, texto, cant in r.productos_nuevos_ignorados[:MAX_EJEMPLOS_REPORTE]:
        print(f"    - {m} ({texto}) en {t}: stock hoy {cant}")


if __name__ == "__main__":
    fecha = date.today()
    if len(sys.argv) == 3:
        dia, mes = sys.argv[1], sys.argv[2]
        fecha = date(fecha.year, int(mes), int(dia))
    resultado = actualizar(fecha)
    imprimir_reporte(resultado)
