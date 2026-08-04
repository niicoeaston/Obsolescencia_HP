# -*- coding: utf-8 -*-
"""
Lee el stock valorizado OFICIAL por zona desde el archivo de referencia
"Base No Estrategicos y obsoletos.xlsx", hoja "Resumen (2)".

No se recalcula nada: se toma tal cual el numero que ya viene en esa hoja
(una tabla dinamica ya armada), en vez de volver a sumar stock x precio
desde el detalle. Esta hoja localiza el bloque dinamicamente (busca el
titulo "Distribucion Por Zona", no asume una fila fija) para no romperse
si el archivo cambia de forma en el futuro.

IMPORTANTE (ver DIAGNOSTICO / README): la hoja "Resumen (2)" tiene DOS
bloques distintos con el mismo titulo "Distribucion Por Zona" (uno cerca
de la columna H, otro cerca de la columna AK) y sus totales NO coinciden
entre si. Se usa el primero (el mas cercano a la columna A, que es el que
esta a la altura de la columna H solicitada) y se deja registrada la
diferencia con el segundo como advertencia -- no se decide cual "esta
bien", eso le corresponde al negocio.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

TITULO_BUSCADO = "distribucion por zona"


@dataclass
class BloqueZona:
    columna_titulo: int
    fila_titulo: int
    filas: list  # [(zona, valor, participacion), ...] sin la fila de total
    total_general: float | None


@dataclass
class ResultadoResumenZona:
    archivo: Path
    hoja: str
    bloques_encontrados: int
    bloque_usado: BloqueZona
    otros_bloques: list = field(default_factory=list)
    advertencias: list = field(default_factory=list)


def _normaliza(s) -> str:
    return str(s).strip().lower() if s is not None else ""


def _leer_bloque(ws, fila_titulo: int, col_titulo: int) -> BloqueZona:
    """
    A partir de la celda del titulo, busca la sub-tabla 'Etiquetas de fila'
    (puede estar unas filas mas abajo, tras las cajas de filtro del pivot)
    y lee zona/valor/participacion hasta la primera fila vacia.
    """
    # Buscar la fila que dice "Etiquetas de fila" en esta misma columna,
    # dentro de las siguientes ~10 filas (las filas de filtro del pivot
    # varian en cantidad, asi que no se asume una posicion fija).
    fila_encabezado = None
    for r in range(fila_titulo, fila_titulo + 12):
        v = ws.cell(row=r, column=col_titulo).value
        if v is not None and "etiqueta" in _normaliza(v):
            fila_encabezado = r
            break
    if fila_encabezado is None:
        raise ValueError(
            f"No se encontro 'Etiquetas de fila' cerca del titulo en fila {fila_titulo}, "
            f"columna {col_titulo}."
        )

    filas = []
    total_general = None
    r = fila_encabezado + 1
    while True:
        zona = ws.cell(row=r, column=col_titulo).value
        valor = ws.cell(row=r, column=col_titulo + 1).value
        participacion = ws.cell(row=r, column=col_titulo + 2).value
        if zona is None or str(zona).strip() == "":
            break
        zona_txt = str(zona).strip()
        if _normaliza(zona_txt) == "total general":
            total_general = float(valor) if isinstance(valor, (int, float)) else None
        else:
            filas.append((zona_txt, float(valor) if isinstance(valor, (int, float)) else None,
                          float(participacion) if isinstance(participacion, (int, float)) else None))
        r += 1

    return BloqueZona(columna_titulo=col_titulo, fila_titulo=fila_titulo,
                       filas=filas, total_general=total_general)


def leer(archivo: Path, hoja: str = "Resumen (2)") -> ResultadoResumenZona:
    wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
    if hoja not in wb.sheetnames:
        raise ValueError(f"El archivo no tiene una hoja '{hoja}'. Hojas: {wb.sheetnames}")
    ws = wb[hoja]

    # Buscar TODAS las celdas cuyo texto sea "Distribucion Por Zona" (puede
    # haber mas de una en la hoja).
    encontrados = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10), max_col=ws.max_column):
        for cell in row:
            if cell.value and TITULO_BUSCADO == _normaliza(cell.value):
                encontrados.append((cell.row, cell.column))

    if not encontrados:
        raise ValueError(f"No se encontro el titulo 'Distribucion Por Zona' en la hoja '{hoja}'.")

    bloques = [_leer_bloque(ws, fr, fc) for fr, fc in encontrados]
    # El solicitado es el que arranca en o mas cerca de la columna H (8).
    bloques.sort(key=lambda b: abs(b.columna_titulo - 8))
    usado, *otros = bloques

    advertencias = []
    if len(bloques) > 1:
        advertencias.append(
            f"La hoja '{hoja}' tiene {len(bloques)} bloques titulados 'Distribucion Por "
            f"Zona' (columnas {[b.columna_titulo for b in bloques]}). Se uso el de la "
            f"columna {usado.columna_titulo} (mas cercano a la columna H pedida)."
        )
        for otro in otros:
            if usado.total_general and otro.total_general:
                dif = usado.total_general - otro.total_general
                if abs(dif) > 1:
                    advertencias.append(
                        f"El bloque de columna {otro.columna_titulo} tiene un total "
                        f"distinto: {otro.total_general:,.0f} vs {usado.total_general:,.0f} "
                        f"usado (diferencia {dif:,.0f}). No se concilia automaticamente."
                    )

    suma_zonas = sum(v for _, v, _ in usado.filas if v is not None)
    if usado.total_general is not None and abs(suma_zonas - usado.total_general) > 1:
        advertencias.append(
            f"La suma de las zonas ({suma_zonas:,.0f}) no coincide exactamente con "
            f"'Total general' ({usado.total_general:,.0f}) en el bloque usado."
        )

    return ResultadoResumenZona(
        archivo=archivo, hoja=hoja, bloques_encontrados=len(bloques),
        bloque_usado=usado, otros_bloques=otros, advertencias=advertencias,
    )


if __name__ == "__main__":
    import sys
    p = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(
        r"C:\Users\NEASTON\OneDrive - DERCO CHILE REPUESTOS SpA\Escritorio\Precios\Acciones "
        r"Comerciales\Venta Obsolecencia\Base No Estrategicos y obsoletos.xlsx"
    )
    res = leer(p)
    print(f"Bloques encontrados: {res.bloques_encontrados}")
    print(f"Bloque usado: columna {res.bloque_usado.columna_titulo}, "
          f"fila titulo {res.bloque_usado.fila_titulo}")
    for zona, val, pct in res.bloque_usado.filas:
        print(f"  {zona:25s} {val:>14,.0f}  {(pct or 0)*100:6.2f}%")
    print(f"  {'Total general':25s} {res.bloque_usado.total_general:>14,.0f}")
    print("\nAdvertencias:")
    for a in res.advertencias:
        print("  -", a)
